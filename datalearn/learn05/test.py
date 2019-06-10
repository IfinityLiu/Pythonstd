import openpyxl
import datetime

# 创建并保存excel文件
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 520
ws.append([1, 2, 3])
ws['A3'] = datetime.datetime.now()
wb.save('demo.xlsx')
